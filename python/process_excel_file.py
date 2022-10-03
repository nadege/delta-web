import sys
import simplekml
from enum import Enum
from openpyxl import load_workbook


def main():

    # Type: 0
    # Date: 1
    # LatLong: 3
    # Altitude: 6

    wb = load_workbook(filename="./data/ActivityReport.xlsx")
    ws = wb[wb.sheetnames[1]]
    print(ws.title)
    data_rows = ws[6:ws.max_row]
    print(len(data_rows))

    kml = simplekml.Kml()

    def get_lat_long(value: str, height: str = None):
        values = [float(height.rstrip(' m'))] + value.split(', ') if height else value.split(', ')
        return tuple(float(val) for val in reversed(values))

    ut = 0
    coords = []
    for row in data_rows:
        if row[0].value == "Check In/OK":
            kml.newpoint(
                coords=[get_lat_long(row[3].value)],
                #description=""
            ),
        elif row[0].value == "Custom Message":
            kml.newpoint(
                coords=[get_lat_long(row[3].value)],
                #description=""
            ),
        elif row[0].value == "Unlimited Track":
            ut = ut + 1
            coords.append(get_lat_long(row[3].value, row[6].value))
        elif row[0].value == "Powered On":
            pass
        else:
            print("unhandled " + row[0].value)

    print("Unlimited Track: " + str(ut))

    line = kml.newlinestring(name="Pathway", coords=coords)
    line.style = simplekml.Style(linestyle=simplekml.LineStyle(width=4, color=simplekml.Color.red))
    kml.save("./data/activity_report.kml")


if __name__ == "__main__":
    sys.exit(main())
